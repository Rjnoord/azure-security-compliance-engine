import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from compliance.scanner import get_resources
from compliance.checks import check_tags, check_region, check_public_ip, check_storage_https, check_nsg_open_ports

resources = get_resources()
findings = []
CHECKS = (
    ("tags", check_tags), ("region", check_region), ("public_ip", check_public_ip),
    ("storage_https", check_storage_https), ("nsg_open_ports", check_nsg_open_ports),
)
for r in resources:
    for check_name, check_fn in CHECKS:
        ok, details = check_fn(r)
        findings.append({
            "name": r.get("name"), "type": r.get("type"),
            "resource_group": r.get("resourceGroup"),
            "check": check_name, "status": "PASS" if ok else "FAIL",
            "details": details,
        })

passed = sum(f["status"] == "PASS" for f in findings)
print(f"Azure Compliance Report\n{'-' * 23}")
print(f"Resources scanned: {len(resources)}")
print(f"Checks passed: {passed}  Checks failed: {len(findings) - passed}")

with open("reports/findings.csv", "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=findings[0].keys())
    writer.writeheader()
    writer.writerows(findings)
print("Findings written to reports/findings.csv")

wb = Workbook()
ws = wb.active
ws.title = "Compliance Findings"
headers = list(findings[0].keys())
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")

fail_fill = PatternFill("solid", fgColor="F8CBAD")
pass_fill = PatternFill("solid", fgColor="C6E0B4")
for finding in findings:
    ws.append(list(finding.values()))
    fill = fail_fill if finding["status"] == "FAIL" else pass_fill
    ws.cell(row=ws.max_row, column=headers.index("status") + 1).fill = fill

for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
wb.save("reports/findings.xlsx")
print("Findings written to reports/findings.xlsx")
